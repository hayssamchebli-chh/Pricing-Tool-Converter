"""Turn a Sheet7/Sheet9-style catalogue extract into the six offer sheets.

The reference workbook is used as a template rather than rebuilt from scratch:
its column widths, number formats, fills and the constants on row 1 of every
offer sheet are exactly what the estimators expect, so the builder only clears
the data band and re-writes rows, footers and the Summary cross-references.
Every derived cell is emitted as a live Excel formula, so the result stays as
editable as the original workbook.
"""

from __future__ import annotations

import io
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from catalog import CatalogItem
from config import (
    CATALOG_FIRST_ROW, CATALOG_HEADERS, CAT_COL, CURRENCY_HEADERS,
    DEFAULT_EUR_FACTOR, DEFAULT_FREIGHT_FACTOR, FACTOR_ROW, FALLBACK_SHEET,
    FIRST_DATA_ROW, LAYOUT_DONORS, SUMMARY_FIRST_ROW, SUMMARY_SHEET, SheetSpec,
    VAT_RATE,
)

SUMMARY_LAST_COL = 4        # Landed usd | Discount | Selling usd | Profit

TEMPLATE_PATH = Path(__file__).with_name("template") / "pricing_tool_template.xlsx"

FOOTER_LABELS = ["Total (USD)", "Discount", "Net Total (USD)", "VAT",
                 "Total Inc. VAT (USD)"]

# Rows priced off ex-works rather than the uploaded landed cost are flagged on
# the Qty cell, in the yellow the reference workbook already used there.
OVER_STOCK_FILL = PatternFill("solid", start_color="FFFFFF00",
                              end_color="FFFFFF00")


# --------------------------------------------------------------------------- #
# routing
# --------------------------------------------------------------------------- #

def prefix_rules(specs):
    """Prefix -> sheet, longest first.

    Longest match matters because the prefixes overlap: ``ABAE-1SVR...`` starts
    with both ``ABA`` and ``ABAE``.
    """
    return sorted(
        ((prefix.strip().upper(), spec.sheet)
         for spec in specs for prefix in spec.prefixes if prefix.strip()),
        key=lambda pair: -len(pair[0]),
    )


def fallback_sheet(specs):
    """Where codes matching no prefix rule land.

    Normally ``FALLBACK_SHEET``, but the routing is editable and that sheet can
    be renamed away or removed, so the last sheet listed stands in for it.
    """
    names = [spec.sheet for spec in specs]
    if FALLBACK_SHEET in names:
        return FALLBACK_SHEET
    return names[-1] if names else FALLBACK_SHEET


def assign_sheets(codes, specs):
    """Map each item code to its target sheet.

    Codes matching no rule go to the fallback sheet so nothing is silently
    dropped from the offer.
    """
    rules = prefix_rules(specs)
    fallback = fallback_sheet(specs)
    assignment = {}
    for code in codes:
        upper = code.upper()
        assignment[code] = next(
            (sheet for prefix, sheet in rules if upper.startswith(prefix)),
            fallback,
        )
    return assignment


@dataclass
class BuildOptions:
    freight_factor: float = DEFAULT_FREIGHT_FACTOR
    eur_factor: float = DEFAULT_EUR_FACTOR
    vat_rate: float = VAT_RATE
    discounts: dict = field(default_factory=dict)   # sheet -> discount fraction


@dataclass
class BuildReport:
    rows_per_sheet: dict = field(default_factory=dict)
    unmatched: list = field(default_factory=list)   # codes with no prefix rule
    catalog_rows: dict = field(default_factory=dict)
    total_rows: dict = field(default_factory=dict)
    fallback: str = ""                              # where unmatched codes went
    added_sheets: list = field(default_factory=list)


# --------------------------------------------------------------------------- #
# low-level sheet helpers
# --------------------------------------------------------------------------- #

def _capture_row_style(worksheet, row, last_col):
    return {col: copy(worksheet.cell(row, col)._style)
            for col in range(1, last_col + 1)}


def _apply_row_style(worksheet, row, styles):
    for col, style in styles.items():
        worksheet.cell(row, col)._style = copy(style)


def _find_footer_row(worksheet, label_col):
    for row in range(FIRST_DATA_ROW, worksheet.max_row + 1):
        if str(worksheet.cell(row, label_col).value or "").strip() == FOOTER_LABELS[0]:
            return row
    return None


def _truncate(worksheet, first_row):
    """Drop every row from ``first_row`` down, values and styling alike."""
    if worksheet.max_row >= first_row:
        worksheet.delete_rows(first_row, worksheet.max_row - first_row + 1)


# --------------------------------------------------------------------------- #
# the Catalogue lookup table
# --------------------------------------------------------------------------- #

def _write_catalog_sheet(worksheet, items):
    last_col = len(CATALOG_HEADERS)
    row_style = _capture_row_style(worksheet, CATALOG_FIRST_ROW, last_col)

    for col, header in enumerate(CATALOG_HEADERS, start=1):
        worksheet.cell(1, col, header)

    row = CATALOG_FIRST_ROW
    for item in items:
        for col, value in enumerate(item.as_row(), start=1):
            worksheet.cell(row, col, value)
        _apply_row_style(worksheet, row, row_style)
        row += 1

    _truncate(worksheet, row)

    if not items:
        # Keep one styled blank row so the sheet's table has a row to cover.
        _apply_row_style(worksheet, CATALOG_FIRST_ROW, row_style)

    last_row = max(row - 1, CATALOG_FIRST_ROW)
    ref = "A1:{}{}".format(get_column_letter(last_col), last_row)
    for table in worksheet.tables.values():
        table.ref = ref
    if worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = ref


# --------------------------------------------------------------------------- #
# offer sheets
# --------------------------------------------------------------------------- #

def _prepare_sheets(workbook, specs):
    """Make the workbook's offer sheets match the specs.

    A spec naming a sheet the template has not got gets one cloned from the
    donor for its layout, so a sheet someone invents carries the same columns,
    widths and number formats as the built-in ones.  Template sheets the specs
    no longer name are dropped, and the tabs finish in the order listed.

    Cloning happens before any removal, so a donor can itself be on the way out.
    """
    for spec in specs:
        if spec.sheet in workbook.sheetnames:
            continue
        sheet = workbook.copy_worksheet(workbook[LAYOUT_DONORS[spec.layout]])
        sheet.title = spec.sheet
        # copy_worksheet carries styles but not the frozen header.
        sheet.freeze_panes = "A3"

    wanted = [spec.sheet for spec in specs]
    keep = set(wanted) | {SUMMARY_SHEET} | {spec.source for spec in specs}
    for name in list(workbook.sheetnames):
        if name not in keep:
            del workbook[name]

    tail = [name for name in workbook.sheetnames if name not in wanted]
    workbook._sheets = [workbook[name] for name in wanted + tail]


def _freight_cell(spec):
    """Absolute address of the freight factor, parked above the landed header."""
    return "${}${}".format(get_column_letter(spec.cols.landed), FACTOR_ROW)


def _landed_cell(spec, row, options):
    """Landed unit cost, chosen by quantity against stock on hand.

    A quantity the stock covers takes the landed cost as uploaded - the
    catalogue value, mirrored into the ``Landed USD`` reference column, which
    is what that stock actually cost to land.  Reaching or exceeding stock
    means the order has to be imported, so the row is priced off the ex-works
    figure instead, grossed up by the freight factor alone for a USD supplier
    and by freight and the EUR conversion for a European one.

    On the ex-works branch the cell stays blank until a price is actually
    keyed, rather than showing a 0 that reads like a costed row.  It fills in
    the moment ``U.P. Ex.`` is entered.

    Freight is read from a cell rather than baked in, so retyping it in the
    sheet reprices every row at once.  The EUR factor stays a literal, fixed
    from the app at build time.
    """
    cols = spec.cols
    qty = "{}{}".format(get_column_letter(cols.qty), row)
    stock = "{}{}".format(get_column_letter(cols.stock), row)
    ex_works = "{}{}".format(get_column_letter(cols.ex_works), row)
    reference = "{}{}".format(get_column_letter(cols.landed_usd), row)
    freight = _freight_cell(spec)
    if spec.currency == "USD":
        gross_up = "{}*{}".format(ex_works, freight)
    else:
        gross_up = "{}*{}*{}".format(ex_works, freight, options.eur_factor)
    return '=IF({q}<{s},{ref},IF({ex}="","",{gross}))'.format(
        q=qty, s=stock, ref=reference, ex=ex_works, gross=gross_up)


def _write_offer_sheet(worksheet, spec, items, quantities, options):
    cols = spec.cols
    source = spec.source

    footer_row = _find_footer_row(worksheet, cols.description)
    data_style = _capture_row_style(worksheet, FIRST_DATA_ROW, cols.last_col)
    if footer_row:
        footer_styles = [_capture_row_style(worksheet, footer_row + offset, cols.last_col)
                         for offset in range(len(FOOTER_LABELS))]
    else:
        footer_styles = [dict(data_style) for _ in FOOTER_LABELS]
    data_height = worksheet.row_dimensions[FIRST_DATA_ROW].height

    _truncate(worksheet, FIRST_DATA_ROW)

    count = len(items)
    first = FIRST_DATA_ROW
    last = first + max(count, 1) - 1        # keep one row so SUM ranges stay valid
    total_row = last + 1
    discount_row = total_row + 1
    net_row = total_row + 2
    vat_row = total_row + 3
    gross_row = total_row + 4

    code_c = get_column_letter(cols.code)
    qty_c = get_column_letter(cols.qty)
    stock_c = get_column_letter(cols.stock)
    price_c = get_column_letter(cols.unit_price)
    landed_c = get_column_letter(cols.landed)
    ex_c = get_column_letter(cols.ex_works)

    # Every landed cell multiplies by this one, so retyping it here reprices
    # the whole sheet.
    worksheet.cell(FACTOR_ROW, cols.landed, options.freight_factor)

    # Name the trading currency in the headers that carry it. A no-op on the
    # built-in sheets, and what stops a cloned sheet inheriting the donor's.
    for field, template in CURRENCY_HEADERS.items():
        worksheet.cell(2, getattr(cols, field), template.format(spec.currency))
    dunit_c = get_column_letter(cols.disc_unit)
    dtotal_c = get_column_letter(cols.disc_total)
    tlanded_c = get_column_letter(cols.total_landed)
    total_c = get_column_letter(cols.total)

    for index, item in enumerate(items):
        row = first + index
        ref = "{}{}".format(code_c, row)

        worksheet.cell(row, cols.num, index + 1)
        worksheet.cell(row, cols.code, item.code)
        worksheet.cell(row, cols.description,
                       "=VLOOKUP({},{}!A:B,2,0)".format(ref, source))
        worksheet.cell(row, cols.qty, quantities.get(item.code, 0))
        worksheet.cell(row, cols.unit_price,
                       "=VLOOKUP({},{}!A:C,3,0)".format(ref, source))
        worksheet.cell(row, cols.landed, _landed_cell(spec, row, options))
        # The three ex-works columns stay blank until a price is keyed into
        # U.P. Ex., rather than showing a row of zeros on everything that is
        # still priced off the catalogue's landed cost.
        worksheet.cell(row, cols.disc_unit,
                       '=IF({ex}{r}=0,"",{ex}{r})'.format(ex=ex_c, r=row))
        worksheet.cell(row, cols.disc,
                       '=IF({ex}{r}=0,"",1-{du}{r}/{ex}{r})'.format(
                           ex=ex_c, du=dunit_c, r=row))
        worksheet.cell(row, cols.disc_total,
                       '=IF({du}{r}="","",{du}{r}*{q}{r})'.format(
                           du=dunit_c, q=qty_c, r=row))
        # An uncosted row leaves U. Landed blank, and blank times a quantity is
        # #VALUE!, so the two cells reading it stay blank in step rather than
        # erroring or claiming a 0.00% margin on a cost nobody has entered.
        worksheet.cell(row, cols.total_landed,
                       '=IF({l}{r}="","",{l}{r}*{q}{r})'.format(
                           l=landed_c, q=qty_c, r=row))
        worksheet.cell(row, cols.total,
                       "={}{}*{}{}".format(price_c, row, qty_c, row))
        worksheet.cell(
            row, cols.margin,
            '=IF({l}{r}="","",'
            'IFERROR(({p}{r}*(1-$D${d})-{l}{r})/({p}{r}*(1-$D${d})),0))'.format(
                p=price_c, l=landed_c, r=row, d=discount_row),
        )
        for col, source_col in (
            (cols.advanced_reserved, CAT_COL["advanced_reserved"]),
            (cols.stock, CAT_COL["stock"]),
            (cols.po_qty, CAT_COL["po_qty"]),
            (cols.po_not_shipped, CAT_COL["po_not_shipped"]),
            (cols.landed_usd, CAT_COL["landed_usd"]),
        ):
            worksheet.cell(row, col, "=VLOOKUP({},{}!A:{},{},0)".format(
                ref, source, get_column_letter(source_col), source_col))

        _apply_row_style(worksheet, row, data_style)
        if data_height:
            worksheet.row_dimensions[row].height = data_height

    if count == 0:
        _apply_row_style(worksheet, first, data_style)

    # Flag the rows the landed formula sends down the ex-works branch - the
    # ones stock cannot cover. A rule rather than a painted fill, so it keeps
    # up as quantities are retyped in Excel. Anchored on the first data row:
    # column absolute, row relative, so Excel walks it down the range.
    worksheet.conditional_formatting.add(
        "{c}{a}:{c}{b}".format(c=qty_c, a=first, b=last),
        FormulaRule(
            formula=["AND(ISNUMBER(${s}{a}),${q}{a}>=${s}{a})".format(
                q=qty_c, s=stock_c, a=first)],
            fill=OVER_STOCK_FILL, stopIfTrue=False,
        ),
    )

    footer = [
        (total_row, {
            cols.disc_total: '=IF(COUNT({c}{a}:{c}{b})=0,"",SUM({c}{a}:{c}{b}))'.format(
                c=dtotal_c, a=first, b=last),
            cols.total_landed: "=SUM({c}{a}:{c}{b})".format(c=tlanded_c, a=first, b=last),
            cols.total: "=SUM({c}{a}:{c}{b})".format(c=total_c, a=first, b=last),
        }),
        (discount_row, {
            cols.qty: options.discounts.get(spec.sheet, 0),
            cols.total: "={c}{t}*$D${d}".format(c=total_c, t=total_row, d=discount_row),
        }),
        (net_row, {
            cols.total: "={c}{t}-{c}{d}".format(c=total_c, t=total_row, d=discount_row),
            cols.margin: (
                "=IFERROR(IF(D{d}=0,({c}{t}-{k}{t})/{c}{t},"
                "({c}{n}-{k}{t})/{c}{n}),0)"
            ).format(c=total_c, k=tlanded_c, t=total_row, d=discount_row, n=net_row),
        }),
        (vat_row, {
            cols.total: "={c}{n}*{v}".format(c=total_c, n=net_row, v=options.vat_rate),
        }),
        (gross_row, {
            cols.total: "={c}{n}+{c}{v}".format(c=total_c, n=net_row, v=vat_row),
        }),
    ]

    for (row, cells), label, styles in zip(footer, FOOTER_LABELS, footer_styles):
        _apply_row_style(worksheet, row, styles)
        worksheet.cell(row, cols.description, label)
        for col, value in cells.items():
            worksheet.cell(row, col, value)

    if worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = "A2:{}{}".format(
            get_column_letter(cols.last_col), last)

    return total_row


def _write_summary(worksheet, specs, total_rows):
    first = SUMMARY_FIRST_ROW
    last = first + len(specs) - 1

    # The template carries one row per built-in sheet plus a bold grand total.
    # Capture both looks up front: adding sheets pushes the total down onto a
    # row that never had styling, and removing them leaves the old one behind.
    data_style = _capture_row_style(worksheet, first, SUMMARY_LAST_COL)
    grand_style = _capture_row_style(worksheet, worksheet.max_row, SUMMARY_LAST_COL)

    for index, spec in enumerate(specs):
        row = first + index
        total_row = total_rows[spec.sheet]
        landed = get_column_letter(spec.cols.total_landed)
        total = get_column_letter(spec.cols.total)
        _apply_row_style(worksheet, row, data_style)
        worksheet.cell(row, 1, "='{}'!{}{}".format(spec.sheet, landed, total_row))
        worksheet.cell(row, 2, None)
        worksheet.cell(row, 3, "='{}'!{}{}".format(spec.sheet, total, total_row + 2))
        worksheet.cell(row, 4, "=IFERROR((C{r}-A{r})/C{r},0)".format(r=row))

    grand = last + 1
    _apply_row_style(worksheet, grand, grand_style)
    worksheet.cell(grand, 1, "=SUM(A{}:A{})".format(first, last))
    worksheet.cell(grand, 2, None)
    worksheet.cell(grand, 3, "=SUM(C{}:C{})".format(first, last))
    worksheet.cell(grand, 4, "=IFERROR((C{r}-A{r})/C{r},0)".format(r=grand))
    _truncate(worksheet, grand + 1)


# --------------------------------------------------------------------------- #
# entry point
# --------------------------------------------------------------------------- #

def build_workbook(items, specs, quantities=None, options=None, template=None):
    """Write the six offer sheets plus Summary and return the workbook bytes."""
    options = options or BuildOptions()
    quantities = quantities or {}
    workbook = openpyxl.load_workbook(template or TEMPLATE_PATH)
    report = BuildReport()

    # Clone before anything is written, so a new sheet copies a pristine donor.
    report.added_sheets = [spec.sheet for spec in specs
                           if spec.sheet not in workbook.sheetnames]
    _prepare_sheets(workbook, specs)

    report.fallback = fallback_sheet(specs)
    assignment = assign_sheets([item.code for item in items], specs)
    rules = prefix_rules(specs)
    report.unmatched = [
        item.code for item in items
        if not any(item.code.upper().startswith(prefix) for prefix, _ in rules)
    ]

    by_sheet = {spec.sheet: [] for spec in specs}
    for item in items:
        by_sheet[assignment[item.code]].append(item)
    for rows in by_sheet.values():
        rows.sort(key=lambda entry: entry.code)

    # Each offer sheet reads description / price / stock from one lookup table,
    # so the catalogue is split the same way the offer sheets are.
    catalog_split = {}
    for spec in specs:
        catalog_split.setdefault(spec.source, []).extend(by_sheet[spec.sheet])
    for source, rows in catalog_split.items():
        rows.sort(key=lambda entry: entry.code)
        _write_catalog_sheet(workbook[source], rows)
        report.catalog_rows[source] = len(rows)

    for spec in specs:
        rows = by_sheet[spec.sheet]
        report.total_rows[spec.sheet] = _write_offer_sheet(
            workbook[spec.sheet], spec, rows, quantities, options)
        report.rows_per_sheet[spec.sheet] = len(rows)

    _write_summary(workbook[SUMMARY_SHEET], specs, report.total_rows)

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream, report
