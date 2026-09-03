"""Reading and normalising the uploaded workbooks.

Two kinds of file are understood:

* a **catalogue** extract - one or more sheets laid out like ``Sheet7`` /
  ``Sheet9`` (item code, description, unit price, stock figures, landed cost);
* an optional **quantity** source - either the raw BOQ or the "sum of qty"
  pivot produced from it, used to fill the ``Qty`` column of the offer sheets.
"""

from __future__ import annotations

import re
from dataclasses import dataclass

import openpyxl

from config import CATALOG_HEADERS

_NUMERIC_FIELDS = (
    "unit_price", "advanced_reserved", "stock",
    "po_qty", "po_not_shipped", "landed_usd",
)

_FIELD_ALIASES = {
    # "no." is deliberately absent: exports that use it for the catalogue code
    # tend to also carry an internal id under that same label, so the column is
    # found by content instead - see _matching_code_column.
    "code": {"item no.1", "item no", "item no.", "item code", "part no.",
             "part no", "part number", "material", "row labels",
             "no.2", "no. 2", "no2"},
    "description": {"description", "desc", "item description"},
    "unit_price": {"unit price", "list price", "price"},
    "advanced_reserved": {"advanced reserved", "advance reserved", "reserved"},
    "stock": {"stock available quantity", "stock available qty", "stock qty",
              "stock", "available quantity"},
    "po_qty": {"po qty", "po quantity", "purchase order qty"},
    "po_not_shipped": {"po not shipped", "po not-shipped", "not shipped"},
    "landed_usd": {"landed usd", "landed (usd)", "landed cost", "landed"},
}

_QTY_ALIASES = {"sum of qty", "sum of quantity", "qty", "quantity", "total qty"}

# What the catalogue itself calls its code column. Used to recognise a
# catalogue sheet and to settle which column holds the codes.
CANONICAL_CODE = "item no.1"


def _norm(value) -> str:
    """Lower-case, collapse whitespace - for tolerant header matching."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip().lower()


def _clean_code(value) -> str:
    """Item codes arrive with stray spaces (pivot tables add a leading one)."""
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).upper()


def _number(value, default=0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[^\d.\-]", "", str(value))
    try:
        return float(text)
    except ValueError:
        return default


@dataclass
class CatalogItem:
    code: str
    description: str = ""
    unit_price: float = 0.0
    advanced_reserved: float = 0.0
    stock: float = 0.0
    po_qty: float = 0.0
    po_not_shipped: float = 0.0
    landed_usd: float = 0.0
    origin: str = ""          # which uploaded sheet the row came from

    def as_row(self) -> list:
        return [
            self.code, self.description, self.unit_price,
            self.advanced_reserved, self.stock, self.po_qty,
            self.po_not_shipped, self.landed_usd,
        ]


def _map_headers(header_row) -> dict:
    """Map ``field name -> 0-based column index`` for one candidate header row."""
    mapping = {}
    for idx, cell in enumerate(header_row):
        label = _norm(cell)
        if not label:
            continue
        for field, aliases in _FIELD_ALIASES.items():
            if label in aliases and field not in mapping:
                mapping[field] = idx
    # The catalogue's own code header outranks any other alias in the row, so a
    # sheet carrying both "Item No.1" and, say, "Part No." keys on the former.
    for idx, cell in enumerate(header_row):
        if _norm(cell) == CANONICAL_CODE:
            mapping["code"] = idx
            break
    return mapping


def _find_header_row(rows, required, scan=12, mode="loose"):
    """Locate the header row: catalogue exports sometimes carry a title above.

    ``mode`` decides how strict a candidate has to be, so a real catalogue can
    be told apart from an offer sheet, whose header row uses the same words:

    * ``canonical`` - the row carries a column headed exactly ``Item No.1``,
      the catalogue's own name for the code column. Offer sheets say "Item
      Code", so this separates them however the columns are arranged.
    * ``contiguous`` - code / description / unit price sit in three
      consecutive columns, the shape of an untouched export.
    * ``loose`` - the three are present anywhere in the row.
    """
    for row_idx, row in enumerate(rows[:scan]):
        mapping = _map_headers(row)
        if not all(field in mapping for field in required):
            continue
        if mode == "canonical":
            if not any(_norm(cell) == CANONICAL_CODE for cell in row):
                continue
        elif mode == "contiguous":
            code = mapping["code"]
            if mapping["description"] != code + 1 or mapping["unit_price"] != code + 2:
                continue
        return row_idx, mapping
    return None, None


def read_catalog(file_like) -> tuple[list[CatalogItem], list[str]]:
    """Return every catalogue row found in an uploaded workbook.

    Any worksheet carrying an item-code + description + unit-price header is
    treated as a catalogue sheet, so a file holding both a ``Sheet7``-style and
    a ``Sheet9``-style tab is read in one pass.  Later duplicates of a code are
    ignored, and the winning row keeps whichever landed cost is non-zero.
    """
    workbook = openpyxl.load_workbook(file_like, data_only=True, read_only=True)
    items: dict[str, CatalogItem] = {}
    skipped: list[str] = []

    sheets = [(s.title, [list(r) for r in s.iter_rows(values_only=True)])
              for s in workbook.worksheets]
    required = ("code", "description", "unit_price")
    # Take the most specific reading the file supports, then apply it to every
    # sheet. Recognising the catalogue by its "Item No.1" header rather than by
    # column adjacency means extra columns can sit between code, description
    # and unit price without the sheet being passed over.
    mode = next(
        (candidate for candidate in ("canonical", "contiguous", "loose")
         if any(_find_header_row(rows, required, mode=candidate)[1]
                for _, rows in sheets if rows)),
        "loose",
    )

    for title, rows in sheets:
        if not rows:
            continue
        header_idx, mapping = _find_header_row(rows, required, mode=mode)
        if mapping is None:
            skipped.append(title)
            continue

        for raw in rows[header_idx + 1:]:
            code = _clean_code(raw[mapping["code"]] if mapping["code"] < len(raw) else None)
            if not code or _norm(code) in {_norm(h) for h in CATALOG_HEADERS}:
                continue
            values = {}
            for field, col in mapping.items():
                cell = raw[col] if col < len(raw) else None
                values[field] = _number(cell) if field in _NUMERIC_FIELDS else cell
            item = CatalogItem(
                code=code,
                description=str(values.get("description") or "").strip(),
                unit_price=values.get("unit_price", 0.0),
                advanced_reserved=values.get("advanced_reserved", 0.0),
                stock=values.get("stock", 0.0),
                po_qty=values.get("po_qty", 0.0),
                po_not_shipped=values.get("po_not_shipped", 0.0),
                landed_usd=values.get("landed_usd", 0.0),
                origin=title,
            )
            existing = items.get(code)
            if existing is None:
                items[code] = item
            elif not existing.landed_usd and item.landed_usd:
                items[code] = item

    workbook.close()
    return list(items.values()), skipped


def _matching_code_column(rows, header_idx, known_codes):
    """Pick the column that actually holds catalogue codes.

    Header names for the code column vary per export - ``No.2`` in one, ``Part
    No.`` in another - and a neighbouring column often carries a different
    identifier entirely, so matching on the name alone picks the wrong one as
    easily as the right one.  Scoring each column against the codes already
    read from the catalogue settles it on the data instead.
    """
    if not known_codes:
        return None
    body = rows[header_idx + 1:]
    width = max((len(row) for row in body), default=0)
    best_col, best_hits = None, 0
    for col in range(width):
        hits = sum(1 for row in body
                   if col < len(row) and _clean_code(row[col]) in known_codes)
        if hits > best_hits:
            best_col, best_hits = col, hits
    return best_col


def read_quantities(file_like, known_codes=None) -> dict[str, dict[str, float]]:
    """Return ``{worksheet: {item code: total qty}}`` for a BOQ workbook.

    Both the raw BOQ and the "sum of qty" pivot built from it usually live in
    the same file and describe the same quantities, so they are kept apart per
    sheet rather than merged - the caller picks one.  Within a sheet a repeated
    code is summed, which is how the same part on several panels adds up.

    Passing ``known_codes`` - the codes read from the catalogue - lets the code
    column be found by content rather than by header name, which is what makes
    an unfamiliar export work without a new alias.
    """
    workbook = openpyxl.load_workbook(file_like, data_only=True, read_only=True)
    known_codes = set(known_codes or ())
    per_sheet: dict[str, dict[str, float]] = {}

    for sheet in workbook.worksheets:
        rows = [list(r) for r in sheet.iter_rows(values_only=True)]
        if not rows:
            continue

        header_idx = qty_col = code_col = None
        for row_idx, row in enumerate(rows[:12]):
            labels = [_norm(c) for c in row]
            candidate_qty = next(
                (i for i, lab in enumerate(labels) if lab in _QTY_ALIASES), None)
            if candidate_qty is None:
                continue
            candidate_code = _matching_code_column(rows, row_idx, known_codes)
            if candidate_code is None:
                candidate_code = _map_headers(row).get("code")
            if candidate_code is not None:
                header_idx, qty_col, code_col = row_idx, candidate_qty, candidate_code
                break
        if header_idx is None:
            continue

        quantities: dict[str, float] = {}
        for raw in rows[header_idx + 1:]:
            code = _clean_code(raw[code_col] if code_col < len(raw) else None)
            if not code or code.startswith("GRANDTOTAL"):
                continue
            qty = _number(raw[qty_col] if qty_col < len(raw) else None)
            if qty:
                quantities[code] = quantities.get(code, 0.0) + qty
        if quantities:
            per_sheet[sheet.title] = quantities

    workbook.close()
    return per_sheet


def best_quantity_sheet(per_sheet: dict[str, dict[str, float]]) -> str | None:
    """Pick the most likely quantity sheet: the one covering the most codes."""
    if not per_sheet:
        return None
    return max(per_sheet, key=lambda name: len(per_sheet[name]))
